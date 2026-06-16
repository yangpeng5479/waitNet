#!/usr/bin/env python3
"""Export Dadibadi Pod website copy mapping to Excel."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# (Page, Section, Original EN, Dadibadi Pod EN, Dadibadi Pod ZH, Dadibadi Pod AR)
ROWS = [
    ("Home", "Page Title", "Glorify Ring | The World's First Christian Smart Ring",
     "Dadibadi Pod | The World's First AI-Powered Story Card Player for Children",
     "Dadibadi Pod | 全球首款儿童 AI 插卡故事机",
     "Dadibadi Pod | أول مشغّل قصص بالبطاقات مدعوم بالذكاء الاصطناعي للأطفال في العالم"),

    ("Home", "Nav - Home", "Home", "Home", "首页", "الرئيسية"),
    ("Home", "Nav - Our Story", "Our Story", "Our Story", "品牌故事", "قصتنا"),
    ("Home", "Nav - App", "Glorify App", "Dadibadi App", "Dadibadi 应用", "تطبيق Dadibadi"),
    ("Home", "Nav CTA", "Join the Waitlist", "Join the Waitlist", "加入等候名单", "انضم إلى قائمة الانتظار"),

    ("Home", "Hero Subtitle", "Glorify Ring", "Dadibadi Pod", "Dadibadi Pod", "Dadibadi Pod"),
    ("Home", "Hero Title Line 1", "The World's First", "The World's First", "全球首款", "الأول في العالم"),
    ("Home", "Hero Title Line 2", "Christian Smart Ring", "AI Story Card Player for Kids",
     "儿童 AI 插卡故事机", "مشغّل قصص بالبطاقات بالذكاء الاصطناعي للأطفال"),
    ("Home", "Hero Description",
     "Glorify Ring is a faith-centered wearable designed to support prayer, reflection, and intentional daily routines.",
     "Dadibadi Pod is a screen-free story companion designed to spark imagination, nurture curiosity, and turn everyday moments into meaningful learning.",
     "Dadibadi Pod 是一款无屏故事伴侣，激发想象力、培养好奇心，让日常每一刻都成为有意义的学习时光。",
     "Dadibadi Pod رفيق قصصي بلا شاشة، صُمّم لإيقاظ الخيال، وتنمية الفضول، وتحويل لحظات اليوم إلى تعلّم هادف ومليء بالمعنى."),
    ("Home", "Hero CTA", "Join the Waitlist", "Join the Waitlist", "加入等候名单", "انضم إلى قائمة الانتظار"),

    ("Home", "Brand Story Title", "The Team Behind the Vision", "The Team Behind the Vision", "愿景背后的团队", "الفريق وراء الرؤية"),
    ("Home", "Brand Story Body",
     "Glorify, The #1 Christian Devotional App has merged with Confidein to bridge this gap between physical life and spiritual journey. Watch our co-founders, Henry and RJ, share the heart behind this union.",
     "Born from a team of parents, educators, and AI innovators, Dadibadi bridges the gap between digital intelligence and the timeless magic of storytelling. Watch our founders share why every child deserves a companion that listens, imagines, and grows with them.",
     "Dadibadi 由父母、教育者与 AI 创新者共同打造，弥合数字智能与永恒故事魅力之间的鸿沟。聆听创始人分享：为何每个孩子都值得拥有一个会倾听、会想象、会陪伴成长的伙伴。",
     "وُلد Dadibadi من فريق من الآباء والمربّين ومبتكري الذكاء الاصطناعي، لسدّ الفجوة بين الذكاء الرقمي وسحر الحكايات الخالد. شاهد مؤسّسينا يشاركونكم لماذا يستحق كل طفل رفيقًا يصغي ويتخيّل وينمو معه."),

    ("Home", "Features Section Title Line 1", "Faith with Focus", "Wonder with Purpose", "有目的的奇趣", "عجائب بقصد"),
    ("Home", "Features Section Title Line 2", "Without Distractions.", "Without Screens.", "远离屏幕。", "بلا شاشات."),
    ("Home", "Features Section Intro",
     "Grow in faith with calm, clarity, and intention. A quiet companion for prayer, reflection, and daily focus.",
     "Grow through stories with calm, curiosity, and intention. A gentle companion for bedtime, playtime, and the quiet moments that shape a child's world.",
     "在平静、好奇与专注中，通过故事陪伴孩子成长。睡前、玩耍时，以及那些塑造孩子世界的安静时刻，它都是温柔的陪伴。",
     "انمو عبر القصص بهدوء وفضول وقصد. رفيق لطيف لوقت النوم وللعب وللحظات الهادئة التي تُشكّل عالم طفلك."),

    ("Home", "Feature 1 - Tag", "Quiet Connection", "Your Voice, Their World", "你的声音，他们的世界", "صوتك، عالمهم"),
    ("Home", "Feature 1 - Title", "Quiet Connection", "Your Voice, Their World", "你的声音，他们的世界", "صوتك، عالمهم"),
    ("Home", "Feature 1 - Subtitle", "Bridging the distance without a single word.",
     "Bridging hearts across every mile — one story at a time.",
     "跨越距离，用心相连——一个故事，一次陪伴。",
     "نربط القلوب عبر كل المسافات — قصة تلو الأخرى."),
    ("Home", "Feature 1 - Bullet 1", "Know when someone needs your prayers",
     "Record your own stories on DIY cards — bedtime tales, lullabies, family memories",
     "在 DIY 卡上录制专属故事——睡前童话、摇篮曲、家庭回忆",
     "سجّل قصصك على بطاقات DIY — حكايات ما قبل النوم، تهاليل، وذكريات عائلية"),
    ("Home", "Feature 1 - Bullet 2", "Silent prayer notifications",
     "Let grandparents and loved ones share their voice, even from afar",
     "让祖辈与亲人远隔千里，也能把声音留在孩子身边",
     "دع الأجداد وأحباءك يشاركون أصواتهم، حتى من بعيد"),
    ("Home", "Feature 1 - Bullet 3", "Shared moments of faith",
     "Create shared moments of wonder that only your family can tell",
     "创造只属于你们家庭的珍贵故事时刻",
     "اصنع لحظات عجب لا ترويها إلا عائلتك"),

    ("Home", "Feature 2 - Tag", "A Gentle Reminder", "Screen-Free Imagination", "无屏想象力", "خيال بلا شاشة"),
    ("Home", "Feature 2 - Title", "A Gentle Reminder", "Screen-Free Imagination", "无屏想象力", "خيال بلا شاشة"),
    ("Home", "Feature 2 - Subtitle", "Faith with Focus, Without Distractions.",
     "Wonder with Purpose, Without Screens.",
     "有目的的奇趣，远离屏幕。",
     "عجائب بقصد، بلا شاشات."),
    ("Home", "Feature 2 - Bullet 1", "Focus back on what matters",
     "Bring focus back to stories, not screens",
     "把注意力还给故事，而非屏幕",
     "أعِد التركيز إلى القصص، لا إلى الشاشات"),
    ("Home", "Feature 2 - Bullet 2", "Mindful daily prompts",
     "Gentle daily story prompts that spark curiosity",
     "每日温柔故事引导，点燃好奇心",
     "محفّزات قصصية يومية لطيفة تُوقظ الفضول"),
    ("Home", "Feature 2 - Bullet 3", "Quiet moments in a busy day",
     "Calm, screen-free moments in a busy childhood",
     "在忙碌的童年里，留一段宁静的无屏时光",
     "لحظات هادئة بلا شاشة في طفولة مزدحمة"),

    ("Home", "Feature 3 - Tag", "Always With You", "Always Listening", "始终倾听", "دائمًا يصغي"),
    ("Home", "Feature 3 - Title", "Always With You", "Always Listening", "始终倾听", "دائمًا يصغي"),
    ("Home", "Feature 3 - Subtitle", "A physical anchor for your most personal relationship.",
     "A trusted companion for your child's growing imagination.",
     "陪伴孩子想象力成长的可靠伙伴。",
     "رفيق موثوق لخيال طفلك المتنامي."),
    ("Home", "Feature 3 - Bullet 1", "Morning to evening reflection",
     "Morning stories to evening lullabies — a rhythm they'll love",
     "从晨间故事到晚间摇篮曲——孩子喜爱的每日节奏",
     "من قصص الصباح إلى تهاليل المساء — إيقاع سيُحبّونه"),
    ("Home", "Feature 3 - Bullet 2", "Devotional rhythm tracking",
     "AI conversations that adapt, remember, and grow with your child",
     "AI 对话随孩子适应、记忆、共同成长",
     "حوارات ذكاء اصطناعي تتكيّف وتتذكّر وتنمو مع طفلك"),
    ("Home", "Feature 3 - Bullet 3", "Your faith companion, always",
     "Their story companion, always ready to listen and wonder",
     "故事伙伴，随时倾听，随时探索",
     "رفيق قصصهم، دائمًا مستعد للإصغاء والتعجّب"),
    ("Home", "Features CTA", "Join the Waitlist", "Join the Waitlist", "加入等候名单", "انضم إلى قائمة الانتظار"),

    ("Home", "Footer - Contact", "Contact us", "Contact us", "联系我们", "تواصل معنا"),
    ("Home", "Footer - Press", "Press", "Press", "媒体合作", "الإعلام"),
    ("Home", "Footer - Partnerships", "Partnerships", "Partnerships", "商务合作", "الشراكات"),
    ("Home", "Footer - Support", "Support", "Support", "帮助支持", "الدعم"),
    ("Home", "Footer - Privacy", "Privacy Policy", "Privacy Policy", "隐私政策", "سياسة الخصوصية"),
    ("Home", "Footer - Terms", "Terms & Conditions", "Terms & Conditions", "条款与条件", "الشروط والأحكام"),
    ("Home", "Footer - Product Link", "Confidein Ring", "Dadibadi DIY Cards", "Dadibadi DIY 卡片", "بطاقات Dadibadi DIY"),
    ("Home", "Footer - Follow", "Follow Us", "Follow Us", "关注我们", "تابعنا"),
    ("Home", "Footer - Download", "Download", "Download", "下载", "تنزيل"),
    ("Home", "Footer - Language", "Language", "Language", "语言", "اللغة"),
    ("Home", "Footer - EN", "EN", "EN", "EN", "EN"),
    ("Home", "Footer - AR", "ES", "AR", "AR", "AR"),

    ("Our Story", "Page Title", "The story behind the union: Glorify & Confidein",
     "The story behind Dadibadi Pod", "Dadibadi Pod 背后的故事", "القصة وراء Dadibadi Pod"),
    ("Our Story", "Opening",
     "When two companies driven by the same mission come together, they don't just combine — they redefine what's possible. This is the moment faith technology becomes whole.",
     "When storytelling meets intelligent technology, something extraordinary happens. This is the moment a child's imagination meets a companion that truly listens.",
     "当讲故事的艺术遇见智能科技，非凡便由此发生。这是孩子想象力与真正会倾听的伙伴相遇的时刻。",
     "حين تلتقي الحكايات بالتقنية الذكية، يحدث شيء استثنائي. هذه اللحظة التي يلتقي فيها خيال الطفل برفيق يصغي حقًا."),
    ("Our Story", "About Label", "The #1 Downloaded Christian Devotional App",
     "Built by Parents, Backed by Educators", "父母打造，教育者背书", "صُنع على يد الآباء، ومدعوم من المربّين"),
    ("Our Story", "About Body",
     "Glorify began with a simple belief: believers deserve a space to connect with God daily. What started as a devotional app has grown into a global platform where millions find stillness, prayer, and scripture.",
     "Dadibadi began with a simple belief: every child deserves stories that feel personal, magical, and made just for them. What started as a question — \"Can technology nurture imagination instead of replacing it?\" — became a screen-free story player loved by families worldwide.",
     "Dadibadi 源于一个朴素的信念：每个孩子都值得拥有专属、神奇、为他们量身定制的故事。从一个问题出发——「科技能否滋养想象力，而非取代它？」——成长为全球家庭喜爱的无屏故事机。",
     "بدأ Dadibadi بإيمان بسيط: كل طفل يستحق قصصًا شخصية وساحرة ومصنوعة خصيصًا له. ما بدأ كسؤال — \"هل يمكن للتقنية أن تغذّي الخيال بدلًا من استبداله؟\" — أصبح مشغّل قصص بلا شاشة تحبه العائلات حول العالم."),
    ("Our Story", "Tags", "BUILT FOR BELIEVERS / BACKED BY VISION / A GLOBAL COMMUNITY",
     "BUILT FOR CHILDREN / BACKED BY SCIENCE / A GLOBAL FAMILY",
     "为孩子而生 / 科学赋能 / 全球家庭", "صُنع للأطفال / مدعوم بالعلم / عائلة عالمية"),
    ("Our Story", "Hardware Label", "Pioneers in Christian Hardware",
     "Pioneers in Child-Safe AI", "儿童安全 AI 先行者", "روّاد الذكاء الاصطناعي الآمن للأطفال"),
    ("Our Story", "Hardware Body",
     "Confidein started with a single question: how do you keep faith present before the phone is even picked up? What started as our debut prayer ring grew and became a movement, with 100,000 people claiming it in 100 days. Today, the same technology team is forging the first ever Christian smart ring.",
     "We started with a single question: how do you give children the wonder of technology — without handing them a screen? Our answer: a tactile story player where every card unlocks a world, and every conversation is safe, warm, and endlessly curious.",
     "我们从一个问题出发：如何让孩子感受科技的美好，却不把屏幕递到他们手中？答案是：一款可触摸的故事机——每张卡片开启一个世界，每次对话都安全、温暖、充满好奇。",
     "بدأنا بسؤال واحد: كيف نمنح الأطفال عجائب التقنية — دون أن نسلّمهم شاشة؟ جوابنا: مشغّل قصص ملموس، كل بطاقة تفتح عالمًا، وكل حوار آمن ودافئ ومليء بالفضول."),
    ("Our Story", "Merger Title", "Redefining Faith Technology", "Redefining Playtime", "重新定义玩耍时光", "إعادة تعريف وقت اللعب"),
    ("Our Story", "Merger Body",
     "When digital and physical converge, a complete ecosystem is created. Millions of believers and Glorify's 25 million existing users now have access to a whole different faith experience. This is not just a merger, this is the moment faith technology becomes whole.",
     "When physical cards and intelligent conversation converge, childhood becomes richer. Insert a card. Hear a story. Ask a question. Watch wonder unfold — naturally, beautifully, without a single screen.",
     "当实体卡片与智能对话相遇，童年变得更加丰盈。插卡、听故事、提问题——看奇趣自然绽放，美好而纯粹，无需一块屏幕。",
     "حين تلتقي البطاقات الملموسة بالحوار الذكي، يصبح الطفولة أغنى. أدخل البطاقة. استمع للقصة. اطرح سؤالًا. وشاهد العجائب تتكشف — بشكل طبيعي وجميل، بلا شاشة واحدة."),
    ("Our Story", "Tomorrow Title", "The Future of Faith is Seamless",
     "The Future of Childhood is Screen-Free", "童年的未来，无屏而生", "مستقبل الطفولة بلا شاشات"),
    ("Our Story", "Tomorrow Body",
     "Faith no longer lives in one place. It lives in the quiet moment you open the app. It lives on your hand throughout the day. It lives in the community that surrounds you.",
     "Wonder no longer lives on a screen. It lives in the card your child chooses. In the question they whisper. In the story only you can record. Dadibadi Pod makes every moment a doorway to imagination.",
     "奇趣不再活在屏幕里。它在孩子挑选的卡片中，在他们轻声提出的问题里，在你录下的独家故事里。Dadibadi Pod 让每一刻，都成为通往想象力的门。",
     "العجائب لم تعد تعيش على شاشة. إنها في البطاقة التي يختارها طفلك. في السؤال الذي يهمس به. في القصة التي لا يمكن لأحد غيرك تسجيلها. Dadibadi Pod يجعل كل لحظة بوابة إلى الخيال."),
    ("Our Story", "Block 1 Title", "Morning to night",
     "Morning to night", "从清晨到夜晚", "من الصباح إلى المساء"),
    ("Our Story", "Block 1 Body",
     "Begin your day with devotion. Carry your faith through every moment. End with reflection and prayer.",
     "Start with a story. End with a lullaby. Every moment in between, a chance to learn and dream.",
     "以故事开启一天，以摇篮曲安然入眠。中间的每一刻，都是学习与梦想的机会。",
     "ابدأ بقصة. انتهِ بتهليلة. وكل لحظة بينهما فرصة للتعلّم والحلم."),
    ("Our Story", "Block 2 Title", "All-day connection",
     "All-day companionship", "全天陪伴", "رفقة طوال اليوم"),
    ("Our Story", "Block 2 Body",
     "The ring becomes your spiritual anchor. Notifications, verses, and community moments arrive when you need them most.",
     "Dadibadi Pod becomes the friend that's always ready to tell one more story.",
     "Dadibadi Pod 成为那个永远愿意「再讲一个故事」的朋友。",
     "يصبح Dadibadi Pod الصديق الذي يكون دائمًا مستعدًا لحكي قصة أخرى."),
    ("Our Story", "Block 3 Title", "One unified vision",
     "One unified vision", "统一的愿景", "رؤية موحّدة"),
    ("Our Story", "Block 3 Body",
     "Digital and physical technology merged into a single, intentional experience. This is what faith technology should be.",
     "Physical cards and AI conversation, designed as one seamless experience for growing minds.",
     "实体卡片与 AI 对话融为一体，为成长中的心灵打造无缝体验。",
     "بطاقات ملموسة وحوار بالذكاء الاصطناعي، في تجربة واحدة سلسة للعقول النامية."),
    ("Our Story", "Join CTA Intro", "Be among the first to experience the future of faith tech.",
     "Be among the first families to experience the future of screen-free storytelling.",
     "成为首批体验无屏讲故事未来的家庭。",
     "كُن من أوائل العائلات التي تجرب مستقبل الحكايات بلا شاشة."),
    ("Our Story", "Join CTA", "Join the Waitlist", "Join the Waitlist", "加入等候名单", "انضم إلى قائمة الانتظار"),

    ("App", "Banner", "Glorify has officially merged with Confidein, introducing the new Glorify Ring. Visit here.",
     "Discover Dadibadi Pod — the screen-free story player with AI conversation and DIY cards. Learn more.",
     "探索 Dadibadi Pod——支持 AI 对话与 DIY 卡片的无屏故事机。了解更多。",
     "اكتشف Dadibadi Pod — مشغّل القصص بلا شاشة مع حوار بالذكاء الاصطناعي وبطاقات DIY. اعرف المزيد."),
    ("App", "Hero Label", "The #1 Christian Devotional App",
     "The Smart Story Companion for Modern Families", "现代家庭的智能故事伴侣", "الرفيق الذكي للقصص للعائلات العصرية"),
    ("App", "Hero Title", "Grow with God. Every day.",
     "Grow Through Stories. Every Day.", "在故事中成长，每一天。", "انمو عبر القصص. كل يوم."),
    ("App", "Hero Sub", "Create your daily devotional habit and strengthen your relationship with God",
     "Create your child's daily story ritual and strengthen the bond between imagination and learning",
     "为孩子建立每日故事仪式，让想象力与学习紧密相连。",
     "اصنع طقس القصص اليومي لطفلك وعزّز الرابط بين الخيال والتعلّم"),
    ("App", "Ticker Words", "Connect / Grow / Find peace / Learn",
     "Imagine / Learn / Play / Wonder", "想象 / 学习 / 玩耍 / 探索", "تخيّل / تعلّم / العب / تأمّل"),
    ("App", "Rhythm Title", "Build Your Daily Rhythm", "Build Your Child's Story Rhythm", "建立孩子的故事节奏", "ابنِ إيقاع قصص طفلك"),
    ("App", "Rhythm 1 Title", "Quote", "Story Card", "故事卡片", "بطاقة قصة"),
    ("App", "Rhythm 1 Body", "Start with an uplifting quote, meant to leave you inspired.",
     "Insert a card and step into a world crafted to spark joy and curiosity.",
     "插入卡片，踏入专为激发快乐与好奇心而打造的世界。",
     "أدخل بطاقة وادخل عالمًا صُمّم لإيقاظ الفرح والفضول."),
    ("App", "Rhythm 2 Title", "Passage", "AI Chat", "AI 对话", "حوار ذكي"),
    ("App", "Rhythm 2 Body", "Build your Bible-reading habit with a daily passage.",
     "Ask questions, explore ideas, and let safe AI conversation deepen every story.",
     "提问、探索，让安全的 AI 对话为每个故事增添深度。",
     "اطرح الأسئلة، واستكشف الأفكار، ودع الحوار الآمن بالذكاء الاصطناعي يعمّق كل قصة."),
    ("App", "Rhythm 3 Title", "Devotional", "DIY Card", "DIY 卡片", "بطاقة DIY"),
    ("App", "Rhythm 3 Body", "Walk through the day's scripture with a guided devotional.",
     "Record your own voice, family tales, and custom content on DIY cards made just for your child.",
     "在专为孩子定制的 DIY 卡上，录制你的声音、家族故事与专属内容。",
     "سجّل صوتك وحكايات العائلة ومحتواك الخاص على بطاقات DIY مصنوعة خصيصًا لطفلك."),
    ("App", "Rhythm 4 Title", "Faith Essentials", "Story Essentials", "故事精选", "أساسيات القصص"),
    ("App", "Rhythm 4 Body", "Strengthen your daily devotional experience with our featured essentials.",
     "Enrich every session with curated story packs, lullabies, and learning adventures.",
     "精选故事包、摇篮曲与学习冒险，丰富每一次聆听体验。",
     "أثرِ كل جلسة بحزم قصص منتقاة وتهاليل ومغامرات تعليمية."),
    ("App", "Section Intro", "Connect with God, no matter where you are.",
     "Connect with stories, no matter where you are.", "无论身在何处，故事与你相伴。", "تواصل مع القصص، أينما كنت."),
    ("App", "Feature 1 Title", "Daily Devotional", "Daily Stories", "每日故事", "قصص يومية"),
    ("App", "Feature 1 Body",
     "Read a daily quote, passage, and devotional, designed to strengthen your faith day by day.",
     "Enjoy fresh stories every day, designed to nurture language, empathy, and a lifelong love of reading.",
     "每日全新故事，培养语言能力、共情力与终身阅读热爱。",
     "استمتع بقصص جديدة كل يوم، مصممة لتنمية اللغة والتعاطف وحب القراءة مدى الحياة."),
    ("App", "Feature 2 Title", "Mindfulness Content", "Calm & Comfort", "安宁与抚慰", "هدوء وراحة"),
    ("App", "Feature 2 Body",
     "Experience God's peace with guided meditations, mental health content, sleep stories, and more.",
     "Wind down with soothing sleep stories, gentle soundscapes, and calming bedtime routines.",
     "舒缓睡前故事、轻柔声景与安宁入睡仪式，帮助孩子安然入眠。",
     "استرخِ مع قصص نوم مهدّئة، وأجواء صوتية لطيفة، وروتين هادئ قبل النوم."),
    ("App", "Feature 3 Title", "Curated Worship Music", "Curated Story Library", "精选故事库", "مكتبة قصص منتقاة"),
    ("App", "Feature 3 Body",
     "From inspiring worship songs to playlists crafted for reflection and sleep, find the perfect sound for your devotional time.",
     "From classic tales to original adventures, find the perfect story for every mood and moment.",
     "从经典童话到原创冒险，为每种心情、每个时刻找到最合适的故事。",
     "من الحكايات الكلاسيكية إلى المغامرات الأصلية، اعثر على القصة المثالية لكل مزاج وكل لحظة."),
    ("App", "Feature 4 Title", "Courses", "Learning Journeys", "学习之旅", "رحلات تعلّم"),
    ("App", "Feature 4 Body",
     "Engage in 5 to 10-day audio journeys that guide you in different areas of life, led by credible authors and theologians.",
     "Explore multi-day story adventures that build vocabulary, values, and confidence — one chapter at a time.",
     "多日连续故事冒险，逐步积累词汇、价值观与自信。",
     "استكشف مغامرات قصصية متعددة الأيام تبني المفردات والقيم والثقة — فصلًا بعد فصل."),
    ("App", "Feature 5 Title", "Glorify Community", "Dadibadi Family", "Dadibadi 家庭社群", "عائلة Dadibadi"),
    ("App", "Feature 5 Body",
     "Don't walk through your faith alone. Connect with like-minded believers and friends with an online community group.",
     "You're not raising curious minds alone. Connect with parents who believe in screen-free, story-rich childhoods.",
     "养育好奇心灵，你并不孤单。与相信无屏、故事丰盈童年的父母一起交流。",
     "لست وحدك في تربية عقول فضولية. تواصل مع آباء يؤمنون بطفولة غنية بالقصص بلا شاشات."),
    ("App", "Testimonial 1",
     "\"I've had this app for a month and haven't missed a day! The streak repair feature has kept me going, and I love it. It's truly changed my life, making it easy to stay in God's Word anywhere with built-in accountability.\"",
     "\"We've had Dadibadi Pod for a month and my daughter asks for it every night. She loves chatting with it about the stories. It's replaced screen time with something I actually feel good about.\"",
     "「用了 Dadibadi Pod 一个月，女儿每晚都要听。她最爱跟它聊故事里的情节。它替换了屏幕时间，这是我真正放心的选择。」",
     "\"امتلكنا Dadibadi Pod منذ شهر وابنتي تطلبه كل ليلة. تحب الحديث معه عن القصص. لقد حلّ محل وقت الشاشة بشيء أشعر بالفعل بالرضا عنه.\""),
    ("App", "Testimonial 2",
     "\"In a hectic season of life, I struggled to make time for God. This app helped me reconnect, calm my anxieties, and find peace in Him. Whether you have minutes or hours, this app is it.\"",
     "\"As a working parent, I struggled to make storytime feel special. Recording my voice on DIY cards changed everything — she hears me even when I'm still at the office.\"",
     "「作为职场父母，我总担心陪读时间不够。在 DIY 卡录下我的声音后，一切变了——即使我还在加班，她也能听到妈妈讲故事。」",
     "\"كأب عامل، كنت أجد صعوبة في جعل وقت القصص مميزًا. تسجيل صوتي على بطاقات DIY غيّر كل شيء — تسمعني حتى عندما أكون لا أزال في المكتب.\""),
    ("App", "Testimonial 3",
     "\"Each morning, this app draws me closer to God, helping to grow my faith, bring encouragement, and provide strength. Whether for study, meditation, or resting in His presence, it allows for spending time with Jesus consistently in a way that works for me!\"",
     "\"Every morning, my son picks a card and his eyes light up. The AI asks the sweetest questions. I've watched his vocabulary and confidence grow in ways no tablet ever gave us.\"",
     "「每天早上，儿子自己选一张卡，眼睛就亮起来。AI 会问最可爱的问题。他的词汇量和自信增长之快，是任何平板都给不了的。」",
     "\"كل صباح، يختار ابني بطاقة وتضيء عيناه. الذكاء الاصطناعي يطرح ألطف الأسئلة. رأيت مفرداته وثقته تنمو بطرق لم تمنحنا إياها أي لوحة رقمية.\""),
    ("App", "Stats Title", "Join Millions in Seeking God", "Join Families Reimagining Storytime", "加入重新定义故事时间的家庭", "انضم إلى العائلات التي تعيد تصور وقت القصص"),
    ("App", "Stat 1", "Over 20 Million Downloads", "Loved by Families Worldwide", "全球家庭的选择", "محبوب من العائلات حول العالم"),
    ("App", "Stat 2", "4.9 Average Rating", "4.9 Average Rating", "4.9 平均评分", "تقييم 4.9 من 5"),
    ("App", "Stat 3", "Over 1,000,000 Ratings", "Parent-Approved & Child-Safe", "家长认可 · 儿童安全", "معتمد من الآباء وآمن للأطفال"),
    ("App", "CTA Primary", "Learn More", "Learn More", "了解更多", "اعرف المزيد"),
    ("App", "CTA Secondary", "Download The App", "Join the Waitlist", "加入等候名单", "انضم إلى قائمة الانتظار"),
    ("App", "Newsletter", "Sign up to receive the latest updates, exclusive content, and our weekly newsletter!",
     "Sign up for launch updates, exclusive story previews, and parenting inspiration.",
     "订阅上线动态、独家故事预览与育儿灵感。",
     "سجّل لتصلك آخر أخبار الإطلاق، ومعاينات حصرية للقصص، وإلهام للتربية."),
    ("App", "Newsletter Field", "Email Address", "Email Address", "电子邮箱", "البريد الإلكتروني"),
    ("App", "Newsletter CTA", "Sign Up", "Sign Up", "订阅", "اشترك"),
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dadibadi Pod Copy"

    headers = [
        "Page", "Section", "Original (Glorify EN)",
        "Dadibadi Pod (EN)", "Dadibadi Pod (ZH)", "Dadibadi Pod (AR)",
    ]
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row_data in enumerate(ROWS, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if col_idx == 6:
                cell.alignment = Alignment(vertical="top", wrap_text=True, readingOrder=2)

    widths = [12, 28, 45, 45, 45, 45]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(ROWS) + 1}"

    out_path = "/Users/hardy/Desktop/py项目/waitNet/Dadibadi_Pod_Website_Copy_EN_AR.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Rows: {len(ROWS)}")


if __name__ == "__main__":
    main()
